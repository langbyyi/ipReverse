#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re
import sys
import socket
import signal
import getopt
import requests
import openpyxl
import time
import threading
from urllib.parse import urlparse
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

HEADER_FILL = PatternFill(start_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
GOV_FILL = PatternFill(start_color='FF0000', fill_type='solid')
GOV_FONT = Font(bold=True, color='FFFFFF')
EDU_FILL = PatternFill(start_color='92D050', fill_type='solid')
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')

BANNER = "\033[93mIP Reverse Domain Lookup | 多源反查 + 合并去重\033[0m"

_rate_lock = threading.Lock()
_last_request_time = 0.0
RATE_LIMIT_SECONDS = 2.0

NO_RESULT_MARKERS = ['no dns a records', 'error', 'no records', 'not found', 'no results']
IP138_TIMEOUT = 5

_source_stats = threading.Lock()
_source_counts = {}

_progress_lock = threading.Lock()
_progress_done = 0
_progress_total = 0

_shutdown = threading.Event()


def _stat_source(name, ok=True, domain_count=0):
    with _source_stats:
        if name not in _source_counts:
            _source_counts[name] = {'ok': 0, 'fail': 0, 'domains': 0}
        if ok:
            _source_counts[name]['ok'] += 1
            _source_counts[name]['domains'] += domain_count
        else:
            _source_counts[name]['fail'] += 1


def _progress_inc():
    global _progress_done
    with _progress_lock:
        _progress_done += 1
        return _progress_done


def rate_limit():
    global _last_request_time
    with _rate_lock:
        now = time.time()
        wait = RATE_LIMIT_SECONDS - (now - _last_request_time)
        if wait > 0:
            time.sleep(wait)
        _last_request_time = time.time()


def resolve_ip(target):
    """从输入中提取hostname并解析为IP，域名和IP都返回IP"""
    try:
        target = target.strip(" '\"")
        if not target.startswith(('http://', 'https://')):
            target = f'http://{target}'
        parsed = urlparse(target)
        hostname = parsed.hostname
        if not hostname:
            return None
        return socket.gethostbyname(hostname)
    except Exception:
        return None


def _safe_get(url, headers=None, timeout=15):
    if _shutdown.is_set():
        return None
    rate_limit()
    if _shutdown.is_set():
        return None
    try:
        session = requests.Session()
        session.trust_env = False
        return session.get(url, headers=headers or {}, timeout=timeout, proxies={'http': None, 'https': None})
    except Exception:
        return None


def fetch_domains_hackertarget(ip):
    domains = []
    try:
        resp = _safe_get(f'https://api.hackertarget.com/reverseiplookup/?q={ip}')
        if _shutdown.is_set():
            return []
        if resp and resp.status_code == 200:
            text = resp.text.strip().lower()
            if any(m in text for m in NO_RESULT_MARKERS):
                _stat_source('hackertarget', ok=True, domain_count=0)
                return []
            for line in resp.text.strip().split('\n'):
                d = line.strip()
                if d and '.' in d and not any(m in d.lower() for m in NO_RESULT_MARKERS):
                    domains.append((d, 'hackertarget'))
            _stat_source('hackertarget', ok=True, domain_count=len(domains))
        else:
            _stat_source('hackertarget', ok=False)
    except Exception:
        _stat_source('hackertarget', ok=False)
    return domains


def fetch_domains_ip138(ip):
    domains = []
    try:
        resp = _safe_get(
            f'https://site.ip138.com/{ip}/',
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Referer': 'https://site.ip138.com/',
            },
            timeout=IP138_TIMEOUT,
        )
        if _shutdown.is_set():
            return []
        if resp and resp.status_code == 200:
            matches = re.findall(r'</span><a href="/([^"]+)/" target="_blank">[^<]+</a></li>', resp.text)
            for m in matches:
                m = m.strip()
                if m:
                    domains.append((m, 'ip138'))
            _stat_source('ip138', ok=True, domain_count=len(domains))
        else:
            _stat_source('ip138', ok=False)
    except Exception:
        _stat_source('ip138', ok=False)
    return domains


def fetch_domains(ip):
    all_domains = []
    for func in [fetch_domains_ip138, fetch_domains_hackertarget]:
        if _shutdown.is_set():
            break
        try:
            result = func(ip)
            if result:
                all_domains.extend(result)
        except Exception:
            continue

    domain_sources = {}
    domain_order = []
    for d, src in all_domains:
        d_lower = d.lower()
        if d_lower not in domain_sources:
            domain_sources[d_lower] = {'display': d, 'sources': []}
            domain_order.append(d_lower)
        if src not in domain_sources[d_lower]['sources']:
            domain_sources[d_lower]['sources'].append(src)

    merged = []
    for d_lower in domain_order[:100]:
        info = domain_sources[d_lower]
        merged.append({
            'domain': info['display'],
            'sources': '+'.join(info['sources']),
        })
    return merged


def check_gov(domain):
    if not domain:
        return False
    lower = domain.lower()
    return any(s in lower for s in ['.gov.cn', '.gov/', '.gov:', '.mil.cn', '.mil/', '.mil:']) or \
           lower.endswith('.gov') or lower.endswith('.mil')


def check_edu(domain):
    if not domain:
        return False
    return '.edu.cn' in domain.lower()


def process_target(target):
    if _shutdown.is_set():
        return (target, None)

    ip = resolve_ip(target)
    if ip is None:
        idx = _progress_inc()
        print(f"\n\033[90m[{idx}/{_progress_total}]\033[0m ► 输入: \033[94m{target}\033[0m")
        print(f"► 失败: 无法解析")
        print("-" * 60)
        return (target, None)

    domain_list = fetch_domains(ip)
    if _shutdown.is_set():
        return (target, None)

    domain_names = [d['domain'] for d in domain_list]
    has_gov = any(check_gov(d) for d in domain_names)
    has_edu = any(check_edu(d) for d in domain_names)

    source_set = set()
    for d in domain_list:
        for s in d['sources'].split('+'):
            source_set.add(s)
    source_tag = '+'.join(sorted(source_set)) if source_set else '-'

    idx = _progress_inc()
    ip_color = "\033[91m" if has_gov else "\033[92m" if domain_list else "\033[90m"
    print(f"\n\033[90m[{idx}/{_progress_total}]\033[0m ► 输入: \033[94m{target}\033[0m")
    print(f"► IP  : {ip_color}{ip}\033[0m")
    print(f"► 域名: {len(domain_list)} 个 \033[90m[来源: {source_tag}]\033[0m", end="")
    if has_gov:
        print(" \033[91m⚠️ 含GOV域名\033[0m", end="")
    if has_edu:
        print(" \033[92m🎓 教育机构\033[0m", end="")
    print()

    if domain_list:
        for d in domain_list[:30]:
            domain = d['domain']
            src = d['sources']
            src_display = f"\033[90m({src})\033[0m" if '+' in src else ""
            if check_gov(domain):
                print(f"  \033[91m⚠️ {domain} (GOV)\033[0m {src_display}".rstrip())
            elif check_edu(domain):
                print(f"  \033[92m🎓 {domain}\033[0m {src_display}".rstrip())
            else:
                print(f"  • {domain} {src_display}".rstrip())
        if len(domain_list) > 30:
            print(f"  ... 还有 {len(domain_list) - 30} 个域名（见输出文件）")
    else:
        print("  (无关联域名)")

    print("-" * 60)
    return (target, {
        "ip": ip,
        "domains": domain_list,
        "has_gov": has_gov,
        "has_edu": has_edu,
        "source_tag": source_tag,
    })


def print_usage():
    print(f"""
Usage: {sys.argv[0]} [OPTIONS]

Options:
  -u, --url <IP/URL/域名>    单个目标反查
  -l, --list <FILE>          批量查询（一行一个，自动跳过#注释行）
  -o, --output <FILE>        输出文件名 (默认: ip_reverse_results.txt)
  -f, --format <txt|xlsx>    输出格式 (默认: txt)
  -t, --threads <NUM>        线程数 (默认: 5)
  -r, --rate <SECONDS>       请求间隔秒数 (默认: 2)
  -h, --help                 显示帮助

Examples:
  {sys.argv[0]} -u 220.181.38.251
  {sys.argv[0]} -u https://example.com
  {sys.argv[0]} -l ips.txt
  {sys.argv[0]} -l ips.txt -o result.xlsx -f xlsx -t 10
  {sys.argv[0]} -l ips.txt -t 3 -r 3

数据来源:
  ip138         site.ip138.com (主力源)
  hackertarget  api.hackertarget.com (备用源)
  结果自动合并去重，多源命中的域名标记所有来源
""")


def export_txt(results, targets_order, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        for target in targets_order:
            data = results.get(target)
            if not data:
                f.write(f"{target}\t\t解析失败\n")
                continue

            domain_names = [d['domain'] for d in data['domains']]
            tags = []
            if data['has_gov']:
                tags.append("GOV")
            if data['has_edu']:
                tags.append("EDU")
            tag_str = ",".join(tags) if tags else "-"

            f.write(f"{target}\t{data['ip']}\t{len(data['domains'])}\t{data['source_tag']}\t{tag_str}\n")
            for d in data['domains']:
                src = d['sources']
                if '+' in src:
                    f.write(f"\t{d['domain']}\t({src})\n")
                else:
                    f.write(f"\t{d['domain']}\n")

    print(f"\n✅ 结果已保存到: {filename}")


def export_xlsx(results, targets_order, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IP反查结果"

    headers = ["原始输入", "IP地址", "关联域名", "域名数", "来源", "标签"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    row_idx = 2
    for target in targets_order:
        data = results.get(target)
        if not data:
            ws.append([target, "", "解析失败", 0, "", ""])
            continue

        domain_list = data['domains']
        domain_lines = []
        for d in domain_list:
            if '+' in d['sources']:
                domain_lines.append(f"{d['domain']} ({d['sources']})")
            else:
                domain_lines.append(d['domain'])

        tags = []
        if data['has_gov']:
            tags.append("GOV")
        if data['has_edu']:
            tags.append("EDU")

        row = [
            target,
            data['ip'],
            "\n".join(domain_lines) if domain_lines else "无结果",
            len(domain_list),
            data.get('source_tag', '-'),
            ",".join(tags) if tags else "-"
        ]
        ws.append(row)

        if data['has_gov']:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = GOV_FILL
                ws.cell(row=row_idx, column=col).font = GOV_FONT
        elif data['has_edu']:
            ws.cell(row=row_idx, column=3).fill = EDU_FILL

        ws.cell(row=row_idx, column=3).alignment = WRAP_ALIGN
        row_idx += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    wb.save(filename)
    print(f"\n✅ 结果已保存到: {filename}")


def main(argv):
    print(f"\n{BANNER}")
    targets = []
    output = ""
    fmt = "txt"
    threads = 5
    global RATE_LIMIT_SECONDS, _source_counts, _progress_done, _progress_total
    _source_counts = {}
    _progress_done = 0
    try:
        opts, args = getopt.getopt(argv, "hu:l:o:f:t:r:", ["help", "url=", "list=", "output=", "format=", "threads=", "rate="])
    except getopt.GetoptError:
        print("参数错误！使用 -h 查看帮助")
        sys.exit(2)

    if not opts:
        print_usage()
        sys.exit(1)

    for opt, arg in opts:
        if opt in ('-h', '--help'):
            print_usage()
            sys.exit()
        elif opt in ("-u", "--url"):
            targets.append(arg)
        elif opt in ("-l", "--list"):
            try:
                with open(arg, 'r') as f:
                    targets.extend(line.strip() for line in f if line.strip() and not line.strip().startswith('#'))
            except FileNotFoundError:
                print(f"文件不存在: {arg}")
                sys.exit(1)
        elif opt in ("-o", "--output"):
            output = arg
        elif opt in ("-f", "--format"):
            fmt = arg.lower()
        elif opt in ("-t", "--threads"):
            threads = int(arg)
        elif opt in ("-r", "--rate"):
            RATE_LIMIT_SECONDS = float(arg)

    if not targets:
        print("请指定目标(-u/-l)")
        sys.exit(1)

    if fmt not in ('txt', 'xlsx'):
        print("格式仅支持 txt 或 xlsx")
        sys.exit(1)

    # 自动推断输出文件名
    if not output:
        output = f"ip_reverse_results.{fmt}"

    # Ctrl+C 优雅退出
    def _signal_handler(sig, frame):
        print(f"\n\n⚠️ 收到中断信号，正在停止...")
        _shutdown.set()

    signal.signal(signal.SIGINT, _signal_handler)

    _progress_total = len(targets)
    source_count = 2
    est_time = len(targets) * source_count * RATE_LIMIT_SECONDS
    print(f"\n🔍 开始处理 {len(targets)} 个目标 (线程: {threads}, 限速: {RATE_LIMIT_SECONDS}s/请求)")
    print(f"📡 数据源: ip138 + hackertarget")
    print(f"📄 输出格式: {fmt} → {output}")
    print(f"⏱️  预计耗时: ~{est_time:.0f}s ({est_time/60:.1f}min)")
    print(f"💡 Ctrl+C 可安全中断\n")

    results = {}
    with ThreadPoolExecutor(max_workers=threads) as executor:
        futures = {executor.submit(process_target, t): t for t in targets}
        try:
            for future in as_completed(futures):
                if _shutdown.is_set():
                    break
                target, data = future.result()
                results[target] = data
        except KeyboardInterrupt:
            _shutdown.set()
            print(f"\n⚠️ 正在取消剩余任务...")

    # 按输入顺序导出
    if fmt == 'xlsx':
        export_xlsx(results, targets, output)
    else:
        export_txt(results, targets, output)

    if _shutdown.is_set():
        interrupted = sum(1 for t in targets if t not in results)
        print(f"\n⚠️ 中断: 已完成 {len(results)}/{len(targets)}, 剩余 {interrupted} 个未处理")

    gov_count = sum(1 for d in results.values() if isinstance(d, dict) and d.get('has_gov'))
    edu_count = sum(1 for d in results.values() if isinstance(d, dict) and d.get('has_edu'))
    other_count = sum(1 for d in results.values() if isinstance(d, dict) and not d.get('has_gov') and not d.get('has_edu'))
    failed = sum(1 for d in results.values() if d is None)
    domains_total = sum(len(d['domains']) for d in results.values() if isinstance(d, dict))
    has_domains = sum(1 for d in results.values() if isinstance(d, dict) and d.get('domains'))

    print(f"\n📊 目标汇总: GOV={gov_count} | EDU={edu_count} | 其他={other_count} | 失败={failed}")
    print(f"📊 域名汇总: 有反查结果={has_domains}/{len(targets)} | 总域名数={domains_total}")

    print(f"\n📡 来源统计:")
    for name in ['ip138', 'hackertarget']:
        s = _source_counts.get(name, {'ok': 0, 'fail': 0, 'domains': 0})
        total = s['ok'] + s['fail']
        rate = f"{s['ok']*100//total}%" if total > 0 else "N/A"
        print(f"  {name:15s} 请求={total} 成功={s['ok']} 失败={s['fail']} 成功率={rate} 贡献域名={s['domains']}")


if __name__ == "__main__":
    main(sys.argv[1:])
